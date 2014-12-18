#this script is responsible for parsing the DrugBank data
__author__ = 'mattdyer'

from optparse import OptionParser
from openpyxl import load_workbook
from openpyxl import Workbook
import xml.etree.ElementTree as ET

#start here when the script is launched
if (__name__ == '__main__'):
    #set up the option parser
    parser = OptionParser()

    #add the options to parse
    parser.add_option('-c', '--compound-xls', dest='compoundxls', help='The master compound spreadsheet')
    parser.add_option('-d', '--drugbank-xml', dest='drugbankxml', help='DrugBank XML File')
    parser.add_option('-o', '--output', dest='outputfile', help='The output file')
    (options, args) = parser.parse_args()

    #load the spreadsheet and parse / store the cas numbers
    wb = load_workbook(filename=options.compoundxls,use_iterators=True)
    ws = wb.get_sheet_by_name('full_list.txt')
    validCAS = {}

    #loop over each row
    for row in ws.iter_rows():
        #see if there are multiple
        if not row[1].value == None:
            for cas in str(row[1].value).split(' '):
                #remove spaces and semicolons
                cas = cas.replace(',', '')
                cas = cas.replace(';', '')
                validCAS[cas] = 1

    #load the drugbank XML file to parse
    drugRecords = ET.parse(options.drugbankxml)
    drugRoot = drugRecords.getroot()

    file = open(options.outputfile, 'w')
    file.write('#cas\tlogP\tlogS\tlogP\tsolubility\tmolecular weight\tmonoisotopic weight\tpsa\trefractivity\tpolarizability\trotateable bond count\th bond acceptor\th bond donor\tpKa acid\tpKa basic\tphysiological change\tnum rings\tbioavailability\trule of five\tghose filter\n')

    for drug in drugRoot.findall('drug'):
        cas = drug.find('cas-number').text

        #only look at the ones in our library for now
        if cas in validCAS:

            #grab the calculated properties
            calculatedProperties = drug.findall('calculated-properties')

            #only look at those with properties
            if len(calculatedProperties) == 1:
                numProperties = 0;
                logPA = 0
                logPC = 0
                logS = 0
                solubility = 0
                molecularWeight = 0
                monoisotopicWeight = 0
                psa = 0
                refractivity = 0
                polarizability = 0
                rotateableBondCount = 0
                hBondAcceptorCount = 0
                hBondDonorCount = 0
                pKaAcid = 0
                pKaBasic = 0
                physiologicalCharge = 0
                numRings = 0
                bioavailability = 0
                ruleOfFive = 0
                ghoseFilter = 0

                for property in calculatedProperties[0].findall('property'):
                    numProperties += 1
                    kind = property.find('kind').text
                    value = property.find('value').text
                    source = property.find('source').text

                    #now we need to figure out the kind and set the value
                    if kind == 'logP' and source == 'ALOGPS':
                        logPA = float(value)
                    elif kind == 'logS':
                        logS = float(value)
                    elif kind == 'logP' and source == 'ChemAxon':
                        logPC = float(value)
                    elif kind == 'Water Solubility':
                        solubility = value.replace(' g/l', '')
                        solubility = float(solubility)
                    elif kind == 'Molecular Weight':
                        molecularWeight = float(value)
                    elif kind == 'Monoisotopic Weight':
                        monoisotopicWeight = float(value)
                    elif kind == 'Polar Surface Area (PSA)':
                        psa = float(value)
                    elif kind == 'Refractivity':
                        refractivity = float(value)
                    elif kind == 'Polarizability':
                        polarizability = float(value)
                    elif kind == 'Rotatable Bond Count':
                        rotateableBondCount = float(value)
                    elif kind == 'H Bond Acceptor Count':
                        hBondAcceptorCount = float(value)
                    elif kind == 'H Bond Donor Count':
                        hBondDonorCount = float(value)
                    elif kind == 'pKa (strongest acidic)':
                        pKaAcid = float(value)
                    elif kind == 'pKa (strongest basic)':
                        pKaBasic = float(value)
                    elif kind == 'Physiological Charge':
                        physiologicalCharge = float(value)
                    elif kind == 'Bioavailability':
                        bioavailability = float(value)
                    elif kind == 'Rule of Five' and value == 'true':
                        ruleOfFive = 1.0
                    elif kind == 'Ghose Filter' and value == 'true':
                        ghoseFilter = 1.0

                #now we print out the line
                if numProperties > 0:
                    file.write('%s\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%f\n' % (cas, logPA, logS, logPC, solubility, molecularWeight, monoisotopicWeight, psa, refractivity, polarizability, rotateableBondCount, hBondAcceptorCount, hBondDonorCount, pKaAcid, pKaBasic, physiologicalCharge, bioavailability, ruleOfFive, ghoseFilter))

    #close the file writer
    file.close()




