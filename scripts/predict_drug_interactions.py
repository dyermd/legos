__author__ = 'mattdyer'

import numpy as np
import pylab as pl

from optparse import OptionParser
from scipy import interp
from sklearn.cross_validation import ShuffleSplit
from sklearn.metrics import roc_curve, auc, confusion_matrix
from sklearn import svm, linear_model
from openpyxl import load_workbook
import random

__author__ = 'mattdyer'

## Run a machine learning model
# @param data The data array object
# @param outputFolder Where to dump results
# @param modelType The type of model to run (Linear SVM, Logistic Regression)
def runModel(data, outputFolder, modelType):
    #store the means for our ROC curve
    meanTPRate = 0.0
    meanFPRate = np.linspace(0, 1, 100)

    #initialize our classifier
    classifier = svm.SVC(kernel='linear', probability=True, random_state=0)

    #see if we wanted logistic regression
    if(modelType == 'Logistic Regression'):
        classifier = linear_model.LogisticRegression(C=1e5, random_state=0)

    #save the confusion matricies for later
    matrices = []

    #start with the subplot
    pl.subplots()

    #now lets analyze the data
    for i, (train, test) in enumerate(data):
        print '\tCross-validation %i' % (i)
        #grab the data sets for training and testing
        xTrain, xTest, yTrain, yTest = X[train], X[test], Y[train], Y[test]

        #train the model
        classifier.fit(xTrain, yTrain)

        #now predict on the hold out
        predictions = classifier.predict(xTest)
        probabilities = classifier.predict_proba(xTest)

        #get the confusion matrix
        matrix = confusion_matrix(yTest, predictions)
        matrices.append(matrix)

        #compute ROC and AUC
        fpr, tpr, thresholds = roc_curve(yTest, probabilities[:, 1])
        meanTPRate += interp(meanFPRate, fpr, tpr)
        meanTPRate[0] = 0.0
        rocAUC = auc(fpr,tpr)

        #now plot it out
        pl.plot(fpr, tpr, lw=1, label='ROC Iter %d (area = %0.2f)' % (i+1, rocAUC))

        #print "P: %s\nA: %s\n" % (predictions, yTest)

    #now plot the random chance line
    pl.plot([0,1], [0,1], '--', color=(0.6,0.6,0.6), label='Random Chance')

    #generate some stats for the mean plot
    meanTPRate /= len(data)
    meanTPRate[-1] = 1.0
    meanAUC = auc(meanFPRate, meanTPRate)

    #plot the average line
    pl.plot(meanFPRate, meanTPRate, 'k--', label='Mean ROC (area = %0.2f)' % (meanAUC), lw=2)

    #add some other plot parameters
    pl.xlim([-0.05, 1.05])
    pl.ylim([-0.05, 1.05])
    pl.xlabel('False Positive Rate')
    pl.ylabel('True Positive Rate')
    pl.title('ROC Curve (%s)' % (modelType))
    pl.legend(loc="lower right")
    pl.savefig('%s/%s.png' % (outputFolder, modelType))

    #plot the confusion matrices
    #for i, (matrix) in enumerate(matrices):
        #pl.subplot(2,3,i + 1)
        #pl.xlim([0, 1.0])
        #pl.ylim([0, 1.0])
        #pl.imshow(matrix, interpolation='nearest', origin='upper')
        #pl.title('Confusion matrix')
        #pl.colorbar()
        #pl.ylabel('Reality')
        #pl.xlabel('Predicted')

#start here when the script is launched
if (__name__ == '__main__'):

    #set up the option parser
    parser = OptionParser()

    #add the options to parse
    parser.add_option('-f', '--folds', dest='folds', help='The cross-fold validations')
    parser.add_option('-p', '--protein', dest='protein', help='Protein feature file')
    parser.add_option('-d', '--drug', dest='drug', help='Drug feature file')
    parser.add_option('-c', '--compound-xls', dest='compoundxls', help='The master compound spreadsheet')
    parser.add_option('-t', '--test', dest='test', help='The fraction of data that goes into the test set (i.e. withheld from training set)')
    parser.add_option('-o', '--output', dest='output', help='The output folder')
    parser.add_option('-m', '--multiplier', dest='multiplier', help='Negative example multiplier')
    (options, args) = parser.parse_args()

    # Load data
    file = open(options.protein, 'r')
    proteins = {}

    for line in file:
        if not line.startswith('#'):
            line = line.rstrip('\n\r')
            tokens = line.split('\t')

            #initiate the array if needed
            if not tokens[0] in proteins:
                proteins[tokens[0]] = []

            #loop over the rest of the file and add the values as floats
            for featureValue in tokens[3:]:
                proteins[tokens[0]].append(float(featureValue))

    file.close()
    file = open(options.drug, 'r')
    drugs = {}

    for line in file:
        if not line.startswith('#'):
            line = line.rstrip('\n\r')
            tokens = line.split('\t')

            #initiate the array if needed
            if not tokens[0] in proteins:
                drugs[tokens[0]] = []

            #loop over the rest of the file and add the values as floats
            #for featureValue in tokens[1:]:
                #drugs[tokens[0]].append(float(featureValue))

    file.close()

    #parse the sheet to load the interactions
    #load the spreadsheet and parse / store the cas numbers
    wb = load_workbook(filename=options.compoundxls,use_iterators=True)
    ws = wb.get_sheet_by_name('full_list.txt')
    data = []
    group = []
    knownInteractions = {}

    #loop over each row
    for row in ws.iter_rows():
        #see if there are multiple
        if not row[1].value == None:
            for cas in str(row[1].value).split(' '):
                #remove spaces and semicolons
                cas = cas.replace(',', '')
                cas = cas.replace(';', '')

                #grab the proteins
                if not row[6].value == None:
                    for protein in str(row[6].value).split('; '):
                        if protein in proteins and cas in drugs:
                            data.append(proteins[protein] + drugs[cas])
                            group.append(1)
                            knownInteractions['%s%s' % (protein, cas)] = 1

    #now we need to generate the negative interactions
    print '%i' % (len(knownInteractions))
    positiveInteractions = len(knownInteractions) * int(options.multiplier)

    while positiveInteractions != 0:
        #generate a random pair
        protein = random.choice(proteins.keys())
        drug = random.choice(drugs.keys())

        #see if this example is already in the list or not
        if not '%s%s' % (protein,drug) in knownInteractions:
            #add it
            knownInteractions['%s%s' % (protein, drug)] = 1
            positiveInteractions -= 1
            data.append(proteins[protein] + drugs[drug])
            group.append(0)

    #we now have our data matrix
    X = np.array(data, dtype=float)
    Y = np.array(group, dtype=int)
    print options.test
    #create the cross validation datasets
    data = ShuffleSplit(len(X), n_iter=int(options.folds), test_size=float(options.test), indices=False, random_state=0)

    #run the SVM classifier
    #print "Starting SVM Test"
    #runModel(data, options.output, 'Linear SVM')

    print "Starting LR Test"
    runModel(data, options.output, 'Logistic Regression')
    #pl.subplot(2, 2, i + 1)

    #plot the graph
    pl.show()




